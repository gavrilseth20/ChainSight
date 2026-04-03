"""
Report Service - Report generation and management
"""
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import uuid
import json
from pathlib import Path
import io

from app.config import settings
from app.core.supabase import supabase_service
from app.core.websocket import ws_manager
from app.services.analysis_service import analysis_service


class ReportService:
    """
    Service for report operations
    """
    
    def __init__(self):
        self.reports_dir = Path(settings.UPLOAD_DIR) / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)
    
    async def generate_report(
        self,
        user_id: str,
        upload_id: str,
        report_type: str,
        report_format: str,
        filters: Optional[Dict] = None
    ) -> Dict:
        """
        Generate a new report
        """
        # Verify upload ownership
        upload = await supabase_service.get_upload_by_id(upload_id)
        if not upload or upload.get("user_id") != user_id:
            raise ValueError("Upload not found")
        
        report_id = str(uuid.uuid4())
        report_name = f"{report_type}_{upload.get('name', 'report')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Create report record
        report = await supabase_service.create_report({
            "id": report_id,
            "user_id": user_id,
            "upload_id": upload_id,
            "name": report_name,
            "type": report_type,
            "format": report_format,
            "status": "generating",
            "filters": filters,
            "created_at": datetime.utcnow().isoformat()
        })
        
        # Generate report asynchronously (simplified - use background task in production)
        try:
            await self._generate_report_content(
                report_id, user_id, upload_id, report_type, report_format, filters
            )
        except Exception as e:
            await supabase_service.update_report(report_id, {
                "status": "failed",
                "error": str(e)
            })
            await ws_manager.broadcast_report_status(
                user_id, report_id, "failed"
            )
        
        return {
            "reportId": report_id,
            "status": "generating",
            "estimatedTime": 30  # seconds
        }
    
    async def _generate_report_content(
        self,
        report_id: str,
        user_id: str,
        upload_id: str,
        report_type: str,
        report_format: str,
        filters: Optional[Dict]
    ):
        """
        Generate the actual report content
        """
        # Get analysis data
        analysis = await analysis_service.get_analysis_results(upload_id, user_id)
        
        if not analysis:
            raise ValueError("No analysis data found")
        
        # Generate based on format
        if report_format == "json":
            content = await self._generate_json_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.json"
            with open(file_path, 'w') as f:
                json.dump(content, f, indent=2, default=str)
        
        elif report_format == "excel":
            content = await self._generate_excel_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.xlsx"
            content.save(file_path)
        
        elif report_format == "pdf":
            content = await self._generate_pdf_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.pdf"
            with open(file_path, 'wb') as f:
                f.write(content)
        
        # Update report record
        file_size = file_path.stat().st_size if file_path.exists() else 0
        
        await supabase_service.update_report(report_id, {
            "status": "completed",
            "file_path": str(file_path),
            "size": file_size,
            "completed_at": datetime.utcnow().isoformat()
        })
        
        # Notify via WebSocket
        await ws_manager.broadcast_report_status(
            user_id, report_id, "completed",
            f"/api/v1/reports/download/{report_id}"
        )
    
    async def _generate_json_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ) -> Dict:
        """
        Generate JSON report
        """
        report = {
            "generatedAt": datetime.utcnow().isoformat(),
            "reportType": report_type,
            "summary": analysis.get("summary", {}),
        }
        
        if report_type in ["compliance", "investigation"]:
            report["patterns"] = analysis.get("patterns", [])
            report["suspiciousAddresses"] = analysis.get("suspiciousAddresses", [])
        
        if report_type == "compliance":
            report["complianceNotes"] = self._generate_compliance_notes(analysis)
        
        return report
    
    async def _generate_excel_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ):
        """
        Generate Excel report
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary = analysis.get("summary", {})
        ws_summary.append(["Metric", "Value"])
        for key, value in summary.items():
            ws_summary.append([key, str(value)])
        
        # Patterns sheet
        if analysis.get("patterns"):
            ws_patterns = wb.create_sheet("Patterns")
            patterns_df = pd.DataFrame(analysis["patterns"])
            
            for r_idx, row in enumerate(dataframe_to_rows(patterns_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_patterns.cell(row=r_idx, column=c_idx, value=str(value) if isinstance(value, list) else value)
        
        # Suspicious Addresses sheet
        if analysis.get("suspiciousAddresses"):
            ws_addresses = wb.create_sheet("Suspicious Addresses")
            addresses_df = pd.DataFrame(analysis["suspiciousAddresses"])
            
            for r_idx, row in enumerate(dataframe_to_rows(addresses_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_addresses.cell(row=r_idx, column=c_idx, value=str(value) if isinstance(value, list) else value)
        
        return wb
    
    async def _generate_pdf_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ) -> bytes:
        """
        Generate PDF report (simplified - use reportlab or weasyprint in production)
        """
        # For hackathon, return a simple text-based PDF
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            elements.append(Paragraph(f"AML Analysis Report - {report_type.title()}", styles['Heading1']))
            elements.append(Spacer(1, 20))
            
            # Summary
            elements.append(Paragraph("Summary", styles['Heading2']))
            summary = analysis.get("summary", {})
            summary_data = [[k, str(v)] for k, v in summary.items()]
            if summary_data:
                t = Table(summary_data, colWidths=[200, 200])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(t)
            
            elements.append(Spacer(1, 20))
            
            # Patterns
            patterns = analysis.get("patterns", [])
            if patterns:
                elements.append(Paragraph("Detected Patterns", styles['Heading2']))
                for p in patterns[:10]:
                    elements.append(Paragraph(
                        f"• {p['type']} ({p['severity']}): {p.get('description', 'N/A')}",
                        styles['Normal']
                    ))
            
            doc.build(elements)
            return buffer.getvalue()
            
        except ImportError:
            # Fallback if reportlab not installed
            return b"PDF generation requires reportlab. Install with: pip install reportlab"
    
    def _generate_compliance_notes(self, analysis: Dict) -> List[str]:
        """
        Generate compliance notes based on analysis
        """
        notes = []
        summary = analysis.get("summary", {})
        
        if summary.get("suspiciousTransactions", 0) > 0:
            notes.append(
                f"ALERT: {summary['suspiciousTransactions']} suspicious transactions detected. "
                "Manual review recommended."
            )
        
        patterns = analysis.get("patterns", [])
        critical_patterns = [p for p in patterns if p.get("severity") == "critical"]
        if critical_patterns:
            notes.append(
                f"CRITICAL: {len(critical_patterns)} critical patterns detected. "
                "Immediate investigation required."
            )
        
        high_risk_addresses = [
            a for a in analysis.get("suspiciousAddresses", [])
            if a.get("riskLevel") in ["critical", "high"]
        ]
        if high_risk_addresses:
            notes.append(
                f"HIGH RISK: {len(high_risk_addresses)} addresses flagged as high/critical risk. "
                "Consider filing SAR."
            )
        
        return notes
    
    async def get_reports(
        self,
        user_id: str,
        upload_id: Optional[str] = None,
        page: int = 1,
        limit: int = 10
    ) -> Tuple[List[Dict], int]:
        """Get reports for a user"""
        reports, total = await supabase_service.get_reports_by_user(
            user_id=user_id,
            upload_id=upload_id,
            page=page,
            limit=limit
        )
        return reports, total or 0
    
    async def get_report_by_id(
        self,
        report_id: str,
        user_id: str
    ) -> Optional[Dict]:
        """Get report by ID"""
        report = await supabase_service.get_report_by_id(report_id, user_id)
        return report
    
    async def get_report_file(
        self,
        report_id: str,
        user_id: str
    ) -> Optional[Dict]:
        """Get report file for download"""
        report = await supabase_service.get_report_by_id(report_id, user_id)
        
        if not report or report.get("status") != "completed":
            return None
        
        file_path = Path(report.get("file_path", ""))
        if not file_path.exists():
            return None
        
        with open(file_path, 'rb') as f:
            content = f.read()
        
        format_type = report.get("format", "json")
        content_types = {
            "pdf": "application/pdf",
            "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "json": "application/json"
        }
        extensions = {"pdf": ".pdf", "excel": ".xlsx", "json": ".json"}
        
        return {
            "content": content,
            "filename": f"{report['name']}{extensions.get(format_type, '')}",
            "media_type": content_types.get(format_type, "application/octet-stream")
        }
    
    async def delete_report(
        self,
        report_id: str,
        user_id: str
    ) -> bool:
        """Delete a report"""
        report = await supabase_service.get_report_by_id(report_id, user_id)
        
        if not report:
            return False
        
        file_path = Path(report.get("file_path", ""))
        if file_path.exists():
            file_path.unlink()
        
        await supabase_service.delete_report(report_id, user_id)
        return True


# Global service instance
report_service = ReportService()
